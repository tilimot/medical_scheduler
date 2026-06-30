import datetime
import csv
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from ortools.sat.python import cp_model

def export_to_rich_excel(dates, doctors, shifts, solver, filename="Planning_Gardes_Ete_2026.xlsx"):
    """Exports the generated schedule to a beautifully formatted Excel workbook with 3 tabs."""
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Styling Definitions
    # -------------------------------------------------------------
    doctor_styles = {
        "Dr 1": {
            "fill": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="0D47A1")
        },
        "Dr 2": {
            "fill": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="1B5E20")
        },
        "Dr 3": {
            "fill": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="4A148C")
        },
        "Dr 4": {
            "fill": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="E65100")
        },
        "Dr 5": {
            "fill": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="880E4F")
        }
    }
    
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='B0BEC5'),
        right=Side(style='thin', color='B0BEC5'),
        top=Side(style='thin', color='B0BEC5'),
        bottom=Side(style='thin', color='B0BEC5')
    )
    
    # -------------------------------------------------------------
    # Tab 1: Calendrier visuel
    # -------------------------------------------------------------
    ws_cal = wb.active
    ws_cal.title = "Calendrier visuel"
    ws_cal.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_cal.merge_cells("A1:G1")
    title_cell = ws_cal["A1"]
    title_cell.value = "Planning de Gardes - Été 2026"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_cal.row_dimensions[1].height = 40
    
    # Subtitle / Info
    ws_cal.merge_cells("A2:G2")
    sub_cell = ws_cal["A2"]
    sub_cell.value = "Vue mensuelle des gardes (Juillet, Août, Septembre 2026)"
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="37474F")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_cal.row_dimensions[2].height = 20
    
    # Set column widths for A to G (1 to 7)
    for c in range(1, 8):
        col_letter = get_column_letter(c)
        ws_cal.column_dimensions[col_letter].width = 18
        
    months = [
        (7, "JUILLET 2026"),
        (8, "AOÛT 2026"),
        (9, "SEPTEMBRE 2026")
    ]
    
    row_idx = 4
    for month_num, month_name in months:
        # Month Header
        ws_cal.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        m_cell = ws_cal.cell(row=row_idx, column=1)
        m_cell.value = month_name
        m_cell.font = Font(name="Calibri", size=14, bold=True, color="37474F")
        m_cell.fill = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")
        m_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_cal.row_dimensions[row_idx].height = 30
        row_idx += 1
        
        # Day Names Header
        headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        ws_cal.row_dimensions[row_idx].height = 20
        for col_idx, h in enumerate(headers):
            cell = ws_cal.cell(row=row_idx, column=col_idx + 1)
            cell.value = h
            cell.font = Font(name="Calibri", size=10, bold=True, color="37474F")
            cell.fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        row_idx += 1
        
        # Weeks grid
        month_weeks = calendar.monthcalendar(2026, month_num)
        
        for week in month_weeks:
            ws_cal.row_dimensions[row_idx].height = 50
            for col_idx, day in enumerate(week):
                cell = ws_cal.cell(row=row_idx, column=col_idx + 1)
                cell.border = thin_border
                
                if day == 0:
                    cell.value = ""
                    cell.fill = gray_fill
                else:
                    day_date = datetime.date(2026, month_num, day)
                    assigned_doc = "Unassigned"
                    for doc in doctors:
                        if solver.Value(shifts[(doc, day_date)]) == 1:
                            assigned_doc = doc
                            break
                    
                    cell.value = f"{day}\n{assigned_doc}"
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    
                    if assigned_doc in doctor_styles:
                        cell.fill = doctor_styles[assigned_doc]["fill"]
                        cell.font = doctor_styles[assigned_doc]["font"]
                    else:
                        cell.font = Font(name="Calibri", size=10, bold=True)
            row_idx += 1
            
        row_idx += 2
        
    # -------------------------------------------------------------
    # Tab 2: Tableau de bord
    # -------------------------------------------------------------
    ws_dashboard = wb.create_sheet(title="Tableau de bord")
    ws_dashboard.views.sheetView[0].showGridLines = True
    
    # Title
    ws_dashboard.merge_cells("A1:D1")
    dash_title = ws_dashboard["A1"]
    dash_title.value = "Tableau de Bord - Statistiques des Gardes"
    dash_title.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    dash_title.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    dash_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_dashboard.row_dimensions[1].height = 35
    
    # Table Headers
    headers = ["Doctor", "Total Shifts", "Weekend Shifts", "Transitions"]
    ws_dashboard.row_dimensions[3].height = 25
    for col_idx, h in enumerate(headers):
        cell = ws_dashboard.cell(row=3, column=col_idx + 1)
        cell.value = h
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    weekend_dates = [day for day in dates if day.weekday() >= 5]
    
    # Populate stats
    row_idx = 4
    for doc in doctors:
        ws_dashboard.row_dimensions[row_idx].height = 20
        shifts_count = sum(solver.Value(shifts[(doc, day)]) for day in dates)
        we_count = sum(solver.Value(shifts[(doc, day)]) for day in weekend_dates)
        
        # Calculate transitions
        trans_count = 0
        for i, day in enumerate(dates):
            if solver.Value(shifts[(doc, day)]) == 1:
                if i == 0 or solver.Value(shifts[(doc, dates[i - 1])]) == 0:
                    trans_count += 1
                    
        c1 = ws_dashboard.cell(row=row_idx, column=1, value=doc)
        c2 = ws_dashboard.cell(row=row_idx, column=2, value=shifts_count)
        c3 = ws_dashboard.cell(row=row_idx, column=3, value=we_count)
        c4 = ws_dashboard.cell(row=row_idx, column=4, value=trans_count)
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c4.alignment = Alignment(horizontal="center", vertical="center")
        
        row_fill = PatternFill(start_color="FFFFFF" if row_idx % 2 == 0 else "F9F9F9", 
                               end_color="FFFFFF" if row_idx % 2 == 0 else "F9F9F9", 
                               fill_type="solid")
        for cell in (c1, c2, c3, c4):
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            
        row_idx += 1
        
    # Total Row
    ws_dashboard.row_dimensions[row_idx].height = 22
    t1 = ws_dashboard.cell(row=row_idx, column=1, value="Total")
    t2 = ws_dashboard.cell(row=row_idx, column=2, value=len(dates))
    t3 = ws_dashboard.cell(row=row_idx, column=3, value=len(weekend_dates))
    t4 = ws_dashboard.cell(row=row_idx, column=4, value=f"=SUM(D4:D{row_idx-1})")
    
    total_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    total_border = Border(
        top=Side(style='thin', color='37474F'),
        bottom=Side(style='double', color='37474F')
    )
    for cell in (t1, t2, t3, t4):
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = total_fill
        cell.border = total_border
        
    t1.alignment = Alignment(horizontal="left", vertical="center")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_dashboard.column_dimensions['A'].width = 15
    ws_dashboard.column_dimensions['B'].width = 15
    ws_dashboard.column_dimensions['C'].width = 18
    ws_dashboard.column_dimensions['D'].width = 15
    
    # Embed Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Distribution de la Charge de Travail"
    chart.y_axis.title = "Number of Shifts"
    chart.x_axis.title = "Doctor"
    
    chart_data = Reference(ws_dashboard, min_col=2, min_row=3, max_col=3, max_row=8)
    chart_cats = Reference(ws_dashboard, min_col=1, min_row=4, max_row=8)
    
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    
    chart.width = 18
    chart.height = 10
    ws_dashboard.add_chart(chart, "F3")
    
    # -------------------------------------------------------------
    # Tab 3: Liste globale
    # -------------------------------------------------------------
    ws_list = wb.create_sheet(title="Liste globale")
    ws_list.views.sheetView[0].showGridLines = True
    
    ws_list.append(["Date", "Day of Week", "Assigned Doctor"])
    for day in dates:
        assigned_doc = "Unassigned"
        for doc in doctors:
            if solver.Value(shifts[(doc, day)]) == 1:
                assigned_doc = doc
                break
        ws_list.append([day.strftime("%Y-%m-%d"), day.strftime("%A"), assigned_doc])
        
    tab = Table(displayName="ScheduleTable", ref=f"A1:C{len(dates)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws_list.add_table(tab)
    
    ws_list.freeze_panes = "A2"
    
    for col in ws_list.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_list.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(filename)

def export_schedule_to_csv(dates, doctors, shifts, solver, filename="medical_schedule_summer_2026.csv"):
    """Exports the generated schedule to a CSV file."""
    with open(filename, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Day of Week", "Assigned Doctor"])
        for day in dates:
            assigned_doc = "Unassigned"
            for doc in doctors:
                if solver.Value(shifts[(doc, day)]) == 1:
                    assigned_doc = doc
                    break
            writer.writerow([day.strftime("%Y-%m-%d"), day.strftime("%A"), assigned_doc])

def get_leave_blocks_and_preceding_weekends(leave_dates, min_leave_len=1):
    """
    Groups a set of leave dates into contiguous blocks of consecutive days,
    and returns the Saturday and Sunday of the weekend immediately preceding
    each block that is at least `min_leave_len` days long.
    """
    if not leave_dates:
        return set()
    
    sorted_leaves = sorted(list(leave_dates))
    blocks = []
    current_block = [sorted_leaves[0]]
    
    for d in sorted_leaves[1:]:
        if (d - current_block[-1]).days == 1:
            current_block.append(d)
        else:
            blocks.append(current_block)
            current_block = [d]
    blocks.append(current_block)
    
    preceding_weekends = set()
    for block in blocks:
        if len(block) < min_leave_len:
            continue
        
        # The first day of the leave block
        first_day = block[0]
        wd = first_day.weekday()  # 0 = Monday, ..., 6 = Sunday
        
        # Calculate the Saturday and Sunday immediately preceding the first day of leave
        # Sunday is (wd + 1) days before the first day of leave
        # Saturday is (wd + 2) days before the first day of leave
        sunday = first_day - datetime.timedelta(days=wd + 1)
        saturday = first_day - datetime.timedelta(days=wd + 2)
        
        preceding_weekends.add(saturday)
        preceding_weekends.add(sunday)
        
    return preceding_weekends

def date_range(start_str, end_str):
    """Generates a list of date objects from start_str to end_str (inclusive)."""
    start = datetime.date.fromisoformat(start_str)
    end = datetime.date.fromisoformat(end_str)
    dates = []
    curr = start
    while curr <= end:
        dates.append(curr)
        curr += datetime.timedelta(days=1)
    return dates

def main():
    # ==========================================
    # 1. PARAMETERS & CONFIGURATION
    # ==========================================
    # Planning horizon: July, August, September 2026 (92 days)
    start_date = datetime.date(2026, 7, 1)
    end_date = datetime.date(2026, 9, 30)
    
    # Generate all days in the horizon
    num_days = (end_date - start_date).days + 1
    dates = [start_date + datetime.timedelta(days=i) for i in range(num_days)]
    
    doctors = ["Dr 1", "Dr 2", "Dr 3", "Dr 4", "Dr 5"]
    
    # Control parameters for constraints
    # Minimum duration (in days) of a leave block to trigger the preceding weekend rule.
    # Set to 1 to strictly enforce the weekend rule for all leaves, including single days.
    # Set to 2 or 3 to avoid restricting weekends for short mid-week leaves (e.g., Dr 4's Wednesday leaves).
    MIN_LEAVE_DURATION_FOR_WEEKEND_RULE = 2 
    
    # Toggle whether the weekend preceding leave rule is a hard or soft constraint
    HARD_WEEKEND_RULE = True
    
    # Weights/Penalties for soft constraints (higher weight = higher priority)
    PENALTY_LEAVE_VIOLATION = 1000
    PENALTY_SPECIAL_REQUEST_VIOLATION = 1000
    PENALTY_WEEKEND_RULE_VIOLATION = 1000  # Only used if HARD_WEEKEND_RULE = False
    
    # Workload Balancing Settings
    ENABLE_HARD_BALANCE = True          # If True, enforces strict minimum and maximum shifts per doctor
    MIN_SHIFTS_PER_DOC = 15            # Minimum shifts any doctor must work
    MAX_SHIFTS_PER_DOC = 22            # Maximum shifts any doctor can work
    
    ENABLE_SOFT_BALANCE = True          # If True, minimizes the deviation from the target number of shifts
    TARGET_SHIFTS_PER_DOC = 18         # Ideal number of shifts (92 days / 5 doctors = 18.4)
    PENALTY_BALANCE_DEVIATION = 50     # Penalty weight per day of deviation from target
    
    # Transition settings to encourage contiguous shifts (minimize fragmentation)
    PENALTY_SHIFT_TRANSITION = 200     # Penalty per shift block start
    
    # Bounded shift block hard constraints
    MAX_CONSECUTIVE_SHIFTS = 7         # Maximum consecutive days a doctor can work
    MIN_REST_DAYS = 3                  # Minimum mandatory rest days after a shift block ends
    
    # Weekend Fairness Settings
    MIN_WEEKEND_SHIFTS = 4             # Minimum weekend shifts any doctor must work
    MAX_WEEKEND_SHIFTS = 7             # Maximum weekend shifts any doctor can work
    TARGET_WEEKEND_SHIFTS = 5          # Ideal number of weekend shifts per doctor (26 total / 5 doctors = 5.2)
    PENALTY_WEEKEND_DEVIATION = 100    # Penalty weight per day of deviation from target
    
    # ==========================================
    # 2. DATA STRUCTURE: LEAVES & SPECIAL REQUESTS
    # ==========================================
    # Define leave intervals for each doctor
    raw_leaves = {
        "Dr 1": [("2026-07-07", "2026-07-21")],
        "Dr 2": [("2026-07-27", "2026-08-14")],
        "Dr 3": [
            ("2026-07-02", "2026-07-07"),
            ("2026-07-16", "2026-07-21"),
            ("2026-08-29", "2026-09-13")
        ],
        "Dr 4": [
            ("2026-07-10", "2026-07-10"),
            ("2026-07-15", "2026-07-15"),
            ("2026-07-22", "2026-07-22"),
            ("2026-07-27", "2026-08-07"),
            ("2026-08-21", "2026-08-21")
        ],
        "Dr 5": [
            ("2026-07-13", "2026-07-17"),
            ("2026-08-03", "2026-08-16")
        ]
    }
    
    # Convert raw leave ranges to sets of date objects
    leaves = {}
    for doc, ranges in raw_leaves.items():
        doc_leaves = set()
        for s_str, e_str in ranges:
            doc_leaves.update(date_range(s_str, e_str))
        leaves[doc] = doc_leaves

    # Special Request: Dr 4 strictly wants to be on call from 06/07 to 12/07
    # Note: Even though he requested leave on 10/07, this special request overrides it.
    dr4_on_call_july_week = set(date_range("2026-07-06", "2026-07-12"))
    
    # We remove 10/07 from Dr 4's leaves because he explicitly wants to work that week
    leaves["Dr 4"] = leaves["Dr 4"] - dr4_on_call_july_week

    # Define all special request dates
    special_on_call = {d: set() for d in doctors}
    special_off_call = {d: set() for d in doctors}
    
    # Dr 4 special requests:
    # 1. On call from 06/07 to 12/07
    special_on_call["Dr 4"].update(date_range("2026-07-06", "2026-07-12"))
    # 2. On call from 24/08 to 30/08
    special_on_call["Dr 4"].update(date_range("2026-08-24", "2026-08-30"))
    # 3. Off call weekend of 22-23/08
    special_off_call["Dr 4"].update(date_range("2026-08-22", "2026-08-23"))

    # Compute preceding weekends for leaves (filtered to scheduling horizon)
    restricted_weekends = {}
    for doc in doctors:
        raw_restricted = get_leave_blocks_and_preceding_weekends(leaves[doc], min_leave_len=MIN_LEAVE_DURATION_FOR_WEEKEND_RULE)
        restricted_weekends[doc] = {d for d in raw_restricted if start_date <= d <= end_date}

    # Identify weekend days (Saturday = 5, Sunday = 6)
    weekend_dates = [day for day in dates if day.weekday() >= 5]

    # ==========================================
    # 3. CP-SAT SOLVER INITIALIZATION
    # ==========================================
    model = cp_model.CpModel()
    
    # Decision Variables: shifts[(doc, day)] = 1 if doctor assigned, 0 otherwise
    shifts = {}
    for doc in doctors:
        for day in dates:
            shifts[(doc, day)] = model.NewBoolVar(f"shift_{doc}_{day.isoformat()}")

    # ==========================================
    # 4. HARD CONSTRAINTS
    # ==========================================
    # Constraint A: Exactly 1 doctor per day
    for day in dates:
        model.Add(sum(shifts[(doc, day)] for doc in doctors) == 1)

    # Constraint B: Weekend immediately preceding leave (Hard if configured)
    for doc in doctors:
        for day in restricted_weekends[doc]:
            if HARD_WEEKEND_RULE:
                model.Add(shifts[(doc, day)] == 0)

    # Constraint C: Maximum consecutive shifts (sliding window)
    for doc in doctors:
        for i in range(num_days - MAX_CONSECUTIVE_SHIFTS):
            model.Add(sum(shifts[(doc, dates[j])] for j in range(i, i + MAX_CONSECUTIVE_SHIFTS + 1)) <= MAX_CONSECUTIVE_SHIFTS)

    # Constraint D: Minimum rest days after a shift block ends
    for doc in doctors:
        for t in range(1, num_days):
            # end_transition represents shifts[t-1] == 1 and shifts[t] == 0
            end_transition = model.NewBoolVar(f"end_trans_{doc}_{dates[t].isoformat()}")
            model.Add(end_transition >= shifts[(doc, dates[t - 1])] - shifts[(doc, dates[t])])
            model.Add(end_transition <= shifts[(doc, dates[t - 1])])
            model.Add(end_transition <= 1 - shifts[(doc, dates[t])])
            
            # If end_transition is True, then shifts[k] must be 0 for k in [t+1, t+MIN_REST_DAYS-1]
            for k in range(t + 1, min(t + MIN_REST_DAYS, num_days)):
                model.AddImplication(end_transition, shifts[(doc, dates[k])].Not())

    # Constraint E: Weekend shifts limits (Hard Constraint)
    for doc in doctors:
        total_weekend_shifts = sum(shifts[(doc, day)] for day in weekend_dates)
        model.Add(total_weekend_shifts >= MIN_WEEKEND_SHIFTS)
        model.Add(total_weekend_shifts <= MAX_WEEKEND_SHIFTS)

    # ==========================================
    # 5. SOFT CONSTRAINTS & OBJECTIVE FUNCTION
    # ==========================================
    penalty_terms = []

    # Soft Constraint 1: Leave days (Penalty if doctor works on leave)
    for doc in doctors:
        for day in leaves[doc]:
            penalty_terms.append(shifts[(doc, day)] * PENALTY_LEAVE_VIOLATION)

    # Soft Constraint 2: Special requests (On-call and Off-call)
    for doc in doctors:
        # Penalty if doctor is NOT on call when requested
        for day in special_on_call[doc]:
            penalty_terms.append((1 - shifts[(doc, day)]) * PENALTY_SPECIAL_REQUEST_VIOLATION)
        # Penalty if doctor IS on call when they requested off
        for day in special_off_call[doc]:
            penalty_terms.append(shifts[(doc, day)] * PENALTY_SPECIAL_REQUEST_VIOLATION)

    # Soft Constraint 3: Preceding weekends (if configured as soft)
    if not HARD_WEEKEND_RULE:
        for doc in doctors:
            for day in restricted_weekends[doc]:
                penalty_terms.append(shifts[(doc, day)] * PENALTY_WEEKEND_RULE_VIOLATION)

    # ==========================================
    # 5.5 WORKLOAD BALANCING CONSTRAINTS
    # ==========================================
    # Hard bounds on shifts per doctor
    if ENABLE_HARD_BALANCE:
        for doc in doctors:
            total_shifts = sum(shifts[(doc, day)] for day in dates)
            model.Add(total_shifts >= MIN_SHIFTS_PER_DOC)
            model.Add(total_shifts <= MAX_SHIFTS_PER_DOC)

    # Soft balancing (minimizing absolute deviation from target)
    if ENABLE_SOFT_BALANCE:
        for doc in doctors:
            total_shifts = sum(shifts[(doc, day)] for day in dates)
            diff = model.NewIntVar(-92, 92, f"diff_{doc}")
            model.Add(diff == total_shifts - TARGET_SHIFTS_PER_DOC)
            
            abs_diff = model.NewIntVar(0, 92, f"abs_diff_{doc}")
            model.AddAbsEquality(abs_diff, diff)
            
            penalty_terms.append(abs_diff * PENALTY_BALANCE_DEVIATION)

    # Weekend soft balancing (minimizing deviation from target weekend shifts)
    for doc in doctors:
        total_weekend_shifts = sum(shifts[(doc, day)] for day in weekend_dates)
        diff_we = model.NewIntVar(-92, 92, f"diff_we_{doc}")
        model.Add(diff_we == total_weekend_shifts - TARGET_WEEKEND_SHIFTS)
        
        abs_diff_we = model.NewIntVar(0, 92, f"abs_diff_we_{doc}")
        model.AddAbsEquality(abs_diff_we, diff_we)
        
        penalty_terms.append(abs_diff_we * PENALTY_WEEKEND_DEVIATION)

    # ==========================================
    # 5.6 TRANSITION CONSTRAINTS (CONTIGUOUS SHIFTS)
    # ==========================================
    # transition_start[(doc, day)] = 1 if doctor doc starts a shift block on day
    transition_start = {}
    for doc in doctors:
        for i, day in enumerate(dates):
            transition_start[(doc, day)] = model.NewBoolVar(f"transition_{doc}_{day.isoformat()}")
            if i == 0:
                # First day: transition start if doctor is on call
                model.Add(transition_start[(doc, day)] == shifts[(doc, day)])
            else:
                prev_day = dates[i - 1]
                # Logic: transition_start = shifts[day] AND NOT shifts[prev_day]
                # We enforce this using three linear inequalities:
                model.Add(transition_start[(doc, day)] >= shifts[(doc, day)] - shifts[(doc, prev_day)])
                model.Add(transition_start[(doc, day)] <= shifts[(doc, day)])
                model.Add(transition_start[(doc, day)] <= 1 - shifts[(doc, prev_day)])
                
            # Add the transition penalty to the objective terms
            penalty_terms.append(transition_start[(doc, day)] * PENALTY_SHIFT_TRANSITION)

    # Define the objective_penalty variable to track and minimize the penalties
    objective_penalty = model.NewIntVar(0, 1000000, "objective_penalty")
    model.Add(objective_penalty == sum(penalty_terms))
    model.Minimize(objective_penalty)

    # ==========================================
    # 6. SOLVE MODEL & PRINT RESULTS
    # ==========================================
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    status = solver.Solve(model)
    
    print("=" * 60)
    print("                MEDICAL SHIFT SCHEDULER")
    print("=" * 60)
    print(f"Solver Status: {solver.StatusName(status)}")
    print(f"Total Objective Penalty: {solver.Value(objective_penalty)}")
    print("-" * 60)
    print(f"Configuration: ")
    print(f"  - Preceding weekend rule is: {'HARD' if HARD_WEEKEND_RULE else 'SOFT'}")
    print(f"  - Min leave duration for weekend rule: {MIN_LEAVE_DURATION_FOR_WEEKEND_RULE} day(s)")
    print(f"  - Hard workload balance: {'ENABLED' if ENABLE_HARD_BALANCE else 'DISABLED'} (Min: {MIN_SHIFTS_PER_DOC}, Max: {MAX_SHIFTS_PER_DOC})")
    print(f"  - Soft workload balance: {'ENABLED' if ENABLE_SOFT_BALANCE else 'DISABLED'} (Target: {TARGET_SHIFTS_PER_DOC}, Penalty: {PENALTY_BALANCE_DEVIATION})")
    print(f"  - Transition penalty: {PENALTY_SHIFT_TRANSITION} (encourages contiguous blocks)")
    print(f"  - Max consecutive shifts (Hard): {MAX_CONSECUTIVE_SHIFTS} days")
    print(f"  - Min rest days (Hard): {MIN_REST_DAYS} days")
    print(f"  - Hard weekend balance: Min: {MIN_WEEKEND_SHIFTS}, Max: {MAX_WEEKEND_SHIFTS}")
    print(f"  - Soft weekend balance: Target: {TARGET_WEEKEND_SHIFTS}, Penalty: {PENALTY_WEEKEND_DEVIATION}")
    print("=" * 60)

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # Print Day-by-Day Schedule
        print("\n--- Day-by-Day Schedule ---")
        current_month = None
        for day in dates:
            if day.month != current_month:
                current_month = day.month
                print(f"\n>>> {day.strftime('%B %Y')} <<<")
            
            assigned_doc = None
            for doc in doctors:
                if solver.Value(shifts[(doc, day)]) == 1:
                    assigned_doc = doc
                    break
            
            # Identify any soft violations to display next to the day
            day_notes = []
            if day in leaves[assigned_doc]:
                day_notes.append("⚠️ Leave day worked")
            if day in restricted_weekends[assigned_doc]:
                day_notes.append("🚨 Preceding weekend worked")
            if day in special_off_call[assigned_doc]:
                day_notes.append("🚫 Requested off-call worked")
            
            # Check for missed on-call requests
            for d in doctors:
                if day in special_on_call[d] and solver.Value(shifts[(d, day)]) == 0:
                    day_notes.append(f"❌ {d} requested to work but was not assigned")
            
            notes_str = f" ({', '.join(day_notes)})" if day_notes else ""
            print(f"  {day.strftime('%Y-%m-%d')} ({day.strftime('%a')}): {assigned_doc}{notes_str}")

        # Print Statistics
        print("\n" + "=" * 60)
        print("                 SCHEDULE STATISTICS")
        print("=" * 60)
        total_assigned_shifts = 0
        total_transitions = 0
        total_weekend_shifts = 0
        for doc in doctors:
            shifts_count = sum(solver.Value(shifts[(doc, day)]) for day in dates)
            transitions_count = sum(solver.Value(transition_start[(doc, day)]) for day in dates)
            weekend_shifts_count = sum(solver.Value(shifts[(doc, day)]) for day in weekend_dates)
            total_assigned_shifts += shifts_count
            total_transitions += transitions_count
            total_weekend_shifts += weekend_shifts_count
            print(f"  {doc}: {shifts_count} shifts, {transitions_count} transitions (blocks), {weekend_shifts_count} weekend shifts")
        print(f"  Total shifts scheduled: {total_assigned_shifts} / {num_days} days (including {total_weekend_shifts} weekend shifts)")
        print(f"  Total transitions: {total_transitions}")
        print("-" * 60)
        
        # Detailed violations breakdown
        print("\nDetailed Violations Breakdown:")
        violations_found = False
        
        for doc in doctors:
            for day in leaves[doc]:
                if solver.Value(shifts[(doc, day)]) == 1:
                    print(f"  [Leave Violation] {doc} worked on leave day: {day.strftime('%Y-%m-%d (%a)')}")
                    violations_found = True
                    
            for day in special_on_call[doc]:
                if solver.Value(shifts[(doc, day)]) == 0:
                    print(f"  [Special Request Violation] {doc} did not work requested on-call day: {day.strftime('%Y-%m-%d (%a)')}")
                    violations_found = True
                    
            for day in special_off_call[doc]:
                if solver.Value(shifts[(doc, day)]) == 1:
                    print(f"  [Special Request Violation] {doc} worked on requested off-call day: {day.strftime('%Y-%m-%d (%a)')}")
                    violations_found = True

            if not HARD_WEEKEND_RULE:
                for day in restricted_weekends[doc]:
                    if solver.Value(shifts[(doc, day)]) == 1:
                        print(f"  [Preceding Weekend Violation] {doc} worked preceding weekend day: {day.strftime('%Y-%m-%d (%a)')}")
                        violations_found = True
                        
        if not violations_found:
            print("  No soft constraints were violated! Optimal schedule found.")
        print("=" * 60)

        # Export schedule to CSV
        csv_filename = "medical_schedule_summer_2026.csv"
        export_schedule_to_csv(dates, doctors, shifts, solver, filename=csv_filename)
        print(f"\nSchedule successfully exported to {csv_filename}")

        # Export schedule to Rich Excel
        excel_filename = "Planning_Gardes_Ete_2026.xlsx"
        export_to_rich_excel(dates, doctors, shifts, solver, filename=excel_filename)
        print(f"Schedule successfully exported to {excel_filename}")
    else:
        print("Could not find a feasible schedule satisfying all hard constraints.")

if __name__ == "__main__":
    main()
